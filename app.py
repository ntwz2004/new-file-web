from flask import Flask,send_file, jsonify, render_template, request, redirect, session, url_for, flash
from flask_login import current_user
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from io import BytesIO

app = Flask(__name__)

# กำหนดค่าฐานข้อมูล
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///user.db'
app.config['SQLALCHEMY_BINDS'] = {
    'patients': 'sqlite:///info.db'  # ฐานข้อมูลสำหรับข้อมูลผู้ป่วย
}
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.secret_key = '1234'  # สำหรับจัดการ session
db = SQLAlchemy(app)

class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(150), unique=True, nullable=False)
    username = db.Column(db.String(150), unique=True, nullable=False)
    password = db.Column(db.String(150), nullable=False)

    def set_password(self, password):
        self.password = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password, password)

class Patient(db.Model):
    __bind_key__ = 'patients'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(150), nullable=False)
    surname = db.Column(db.String(150), nullable=False)
    dental_num = db.Column(db.String(50), nullable=False)
    diagnosis = db.Column(db.Text, nullable=True)  # เปลี่ยนเป็น Text เพื่อรองรับข้อมูลหลายรายการ
    icd10 = db.Column(db.Text, nullable=True)      # เปลี่ยนเป็น Text เช่นกัน
    visit_type = db.Column(db.String(50), nullable=False)
    date = db.Column(db.Date, nullable=False)
    created_by = db.Column(db.String(150), nullable=True)

    def add_diagnosis(self, diagnosis, icd_code):
        if self.diagnosis:
            self.diagnosis += f",{diagnosis}"
            self.icd10 += f",{icd_code}"
        else:
            self.diagnosis = diagnosis
            self.icd10 = icd_code


@app.route('/')
def index():
    return render_template('index.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email_or_username = request.form['email_or_username']
        password = request.form['password']
        user = User.query.filter((User.email == email_or_username) | (User.username == email_or_username)).first()
        if user and user.check_password(password):
            session['username'] = user.username
            return jsonify(success=True , message="Login successful")  # ส่งข้อความสำเร็จกลับไป
        else:
            return jsonify(success=False, message="The username or password is incorrect")
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.pop('username', None)  # ลบข้อมูล session ของ username
    return redirect(url_for('index'))

@app.context_processor
def inject_user():
    return dict(current_user=session.get('username'))

@app.route('/reg', methods=['GET', 'POST'])
def reg():
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        username = request.form.get('username')

        existing_user = User.query.filter((User .email == email) | (User .username == username)).first()
        if existing_user:
            return {"success": False, "message": "This email or username already exists"}, 400

        new_user = User(email=email, username=username)
        new_user.set_password(password)
        db.session.add(new_user)
        db.session.commit()
        
        return jsonify(success=True, message="Successfully applied for membership")  # ส่งข้อความสำเร็จกลับไป
    return render_template('reg.html')

@app.route('/main')
def main():
    patients = Patient.query.all()
    return render_template('main.html', patients=patients)

@app.route('/search')
def search():
    return render_template('search.html')

# ฟังก์ชันใหม่สำหรับค้นหาข้อมูล
@app.route('/search_results', methods=['POST'])
def search_results():
    data = request.get_json()
    filter_type = data.get("filterType")
    filter_value = data.get("filterValue", "").lower()

    query = db.session.query(Patient)

    if filter_type == "name":
        query = query.filter(Patient.name.ilike(f"%{filter_value}%"))
    elif filter_type == "surname":
        query = query.filter(Patient.surname.ilike(f"%{filter_value}%"))
    elif filter_type == "dental_number":
        query = query.filter(Patient.dental_num.ilike(f"%{filter_value}%"))
    elif filter_type == "diagnosis":
        query = query.filter(Patient.diagnosis.ilike(f"%{filter_value}%"))
    elif filter_type == "icd_10":
        query = query.filter(Patient.icd10.ilike(f"%{filter_value}%"))
    elif filter_type == "type_of_visit":
        query = query.filter(Patient.visit_type.ilike(f"%{filter_value}%"))
    elif filter_type == "date":
        try:
            search_date = datetime.strptime(filter_value, "%Y-%m-%d").date()
            query = query.filter(Patient.date == search_date)
        except ValueError:
            return jsonify([])

    results = []
    for patient in query.all():
        diagnoses = patient.diagnosis.split(",") if patient.diagnosis else ["-"]
        icd10_codes = patient.icd10.split(",") if patient.icd10 else ["-"]

        diagnosis_text = "<br>".join(diagnoses).strip() if diagnoses else "-"
        icd_text = "<br>".join(icd10_codes).strip() if icd10_codes else "-"

        results.append({
            "id": patient.id,  # เพิ่ม ID
            "name": patient.name,
            "surname": patient.surname,
            "dental_number": patient.dental_num,
            "diagnosis": diagnosis_text,
            "icd_10": icd_text,
            "type_of_visit": patient.visit_type,
            "date": patient.date.strftime("%Y-%m-%d") if patient.date else "-"
        })

    return jsonify(results)


@app.route('/add', methods=['GET', 'POST'])
def add():
    if request.method == 'POST':
        name = request.form.get('name')
        surname = request.form.get('surname')
        dental_num = request.form.get('dental_num')
        visit_type = request.form.get('visit_type')
        date_str = request.form.get('date')
        date = datetime.strptime(date_str, '%Y-%m-%d').date()
        
        # เก็บชื่อผู้ใช้ที่เพิ่มข้อมูล
        created_by = session.get('username')  # ดึงข้อมูล username จาก session

        # เพิ่มผู้ป่วยใหม่และบันทึกชื่อผู้ใช้ที่ทำการเพิ่มข้อมูล
        new_patient = Patient(
            name=name,
            surname=surname,
            dental_num=dental_num,
            visit_type=visit_type,
            date=date,
            created_by=created_by
        )
        db.session.add(new_patient)
        db.session.commit()

        # เพิ่มการวินิจฉัยหลายรายการ
        diagnoses = request.form.getlist('diagnosis[]')
        icd10_codes = request.form.getlist('icd10[]')
        
        for diagnosis, icd10 in zip(diagnoses, icd10_codes):
            if diagnosis or icd10:
                new_patient.add_diagnosis(diagnosis, icd10)

        db.session.commit()
        flash('เพิ่มข้อมูลผู้ป่วยสำเร็จ!')
        return jsonify({"success": True})

    return render_template('add.html')


@app.route('/get_patient_data/<int:patient_id>')
def get_patient_data(patient_id):
    patient = Patient.query.get_or_404(patient_id)
    data = {
        "id": patient.id,
        "name": patient.name,
        "surname": patient.surname,
        "dental_number": patient.dental_num,
        "diagnosis": patient.diagnosis,
        "icd10": patient.icd10,
        "type_of_visit": patient.visit_type,
        "date": patient.date.strftime("%Y-%m-%d") if patient.date else "-",
        "created_by": patient.created_by
    }
    return jsonify(data)

@app.route('/edit_patient/<int:patient_id>', methods=['POST'])
def edit_patient(patient_id):
    data = request.get_json()
    patient = Patient.query.get_or_404(patient_id)
    patient.name = data.get('name')
    patient.surname = data.get('surname')
    patient.dental_num = data.get('dental_number')
    patient.diagnosis = data.get('diagnosis')
    patient.icd10 = data.get('icd10')
    patient.visit_type = data.get('type_of_visit')
    patient.date = datetime.strptime(data.get('date'), "%Y-%m-%d") if data.get('date') else None

    db.session.commit()
    return jsonify({"success": True})

@app.route('/delete_patient/<int:patient_id>', methods=['DELETE'])
def delete_patient(patient_id):
    patient = Patient.query.get(patient_id)
    if patient:
        db.session.delete(patient)
        db.session.commit()
        return jsonify(success=True, message="Data has been successfully deleted")
    return jsonify(success=False, message="The data you want to delete was not found")
    
@app.route('/add_visit/<int:patient_id>', methods=['POST'])
def add_visit(patient_id):
    data = request.get_json()

    diagnosis = data.get('diagnosis','')
    icd10 = data.get('icd10','')

    # Save the diagnosis and icd10 values as strings in the database
    new_visit = Patient(
        name=data['name'],
        surname=data['surname'],
        dental_num=data['dental_number'],
        diagnosis=diagnosis,  # Save as a single string
        icd10=icd10,  # Save as a single string
        visit_type="Follow up",
        date=datetime.strptime(data['date'], "%Y-%m-%d").date(),
        created_by=session.get('username')
    )
    
    db.session.add(new_visit)
    
    try:
        db.session.commit()
        return jsonify({"success": True, "message": "New visit added successfully"})
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "message": str(e)})
    
def clean_multiline_text(text):
    # Ensure text is a string and replace HTML line breaks with newline
    if text is None:
        return ""
    return str(text).replace('<br>', '\n').strip()

@app.route('/export_to_excel', methods=['POST'])
def export_to_excel():
    data = request.get_json()
    
    # Create a new workbook and select the active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Patient Data"
    
    # Define styles
    header_font = Font(bold=True, color="000000")
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # Add headers
    headers = [
        "Dental Number", "Name", "Surname", "Diagnosis", 
        "ICD-10", "Type of Visit", "Date"
    ]
    
    # Write headers with styling
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Adjust column width
    column_widths = [15, 15, 15, 25, 25, 15, 12]  # Increased ICD-10 column width
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width
    
    # Add data rows
    for row_num, item in enumerate(data, 2):
        # Clean all text fields
        dental_number = clean_multiline_text(item.get('dental_number', ''))
        name = clean_multiline_text(item.get('name', ''))
        surname = clean_multiline_text(item.get('surname', ''))
        diagnosis = clean_multiline_text(item.get('diagnosis', ''))
        icd_10 = clean_multiline_text(item.get('icd_10', ''))
        type_of_visit = clean_multiline_text(item.get('type_of_visit', ''))
        date = clean_multiline_text(item.get('date', ''))
        
        # Write data to cells
        ws.cell(row=row_num, column=1, value=dental_number)
        ws.cell(row=row_num, column=2, value=name)
        ws.cell(row=row_num, column=3, value=surname)
        ws.cell(row=row_num, column=4, value=diagnosis)
        ws.cell(row=row_num, column=5, value=icd_10)
        ws.cell(row=row_num, column=6, value=type_of_visit)
        ws.cell(row=row_num, column=7, value=date)
        
        # Apply styling to each cell
        for col in range(1, 8):
            cell = ws.cell(row=row_num, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(
                horizontal='left', 
                vertical='center', 
                wrap_text=True
            )
    
    # Enable auto-filtering
    ws.auto_filter.ref = ws.dimensions
    
    # Save to a BytesIO object
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Send the file
    return send_file(
        excel_file, 
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, 
        download_name='patient_data.xlsx'
    ) 

# สร้างตารางฐานข้อมูลเมื่อเริ่มต้น
with app.app_context():
    db.create_all()

if __name__ == '__main__':
    app.run(debug=True)